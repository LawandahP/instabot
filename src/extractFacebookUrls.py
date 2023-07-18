from bs4 import BeautifulSoup
from colorama import Fore
import openpyxl
import requests


import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class UrlBot():
    def __init__(self):
        self.filename = '/home/pablo/Projects/instabot/src/Data-Pull-Template.xlsx'



    def progress_bar(self, progress, total):
        percent = 100 * (progress / float(total))
        bar = '◼︎' * int(percent) + '' * (100 - int(percent))
        print(Fore.GREEN + f"\r|{bar}| {percent:.2f}%", end="\r")
        
    def getCompanyUrls(self):
        
        excel_file = self.filename
        workbook = openpyxl.load_workbook(excel_file)

        # Select the desired worksheet
        worksheet = workbook.active  

        # Specify the column letter (e.g., 'A' for the first column)
        column_letter = 1  # Replace 'A' with the desired column letter

        # Extract data from the specified column
        column_data = []
        
        for cell in worksheet.iter_rows(min_row=2, min_col=column_letter, max_col=column_letter, values_only=True):
            column_data.append(cell[0])

            # self.progress_bar(iter, len(column_data))
            if len(column_data) == 5:
                break

        return column_data
    
    def getSocialMediaUrls(self, links=[], media_url=str, site=str):

        media_links = []
        
        iter = 0
        print(f"Getting {site} links...")

        self.progress_bar(iter, len(links))
        for link in links:
            iter += 1
            response = requests.get(f"https://{link}", verify=False)

            # Parse the HTML content using BeautifulSoup
            soup = BeautifulSoup(response.content, 'html.parser')

            # Find the link
            link = "None"

            for link in soup.find_all('a', href=lambda href: href and media_url in href):
                href = link.get('href')

                self.progress_bar(iter, len(links))
                if href and f'{site}.com' in href:
                    link = href
                    break
                else:
                    break
            media_links.append(link)
        return media_links


    def saveToExcel(self, start_column=str, profiles=[]):
        # Open the Excel file
        excel_file = self.filename
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active

        start_column = start_column  # Replace with the desired starting column
        start_row = 2  # Replace with the desired starting row


        for i, value in enumerate(profiles, start=start_row):
            worksheet.cell(row=i, column=openpyxl.utils.column_index_from_string(start_column), value=value)

        # Save the changes
        workbook.save(excel_file)


bot = UrlBot()

urls = bot.getCompanyUrls()


# facebook
facebook_links = bot.getSocialMediaUrls(links=urls, media_url="https://www.facebook.com/", site='facebook')
save_facebook_links = bot.saveToExcel(start_column='J', profiles=facebook_links)

# instagram
instagram = bot.getSocialMediaUrls(links=urls, media_url="https://www.instagram.com/", site='instagram')
save_ig_links = bot.saveToExcel(start_column='K', profiles=instagram)

# instagram
youtube = bot.getSocialMediaUrls(links=urls, media_url="https://www.youtube.com/", site='youtube')
save_youtube_links = bot.saveToExcel(start_column='L', profiles=youtube)

# instagram
tiktok = bot.getSocialMediaUrls(links=urls, media_url="https://www.tiktok.com/", site='tiktok')
save_tiktok_links = bot.saveToExcel(start_column='M', profiles=tiktok)

