import pickle
import time
import os
from typing import Any
from bs4 import BeautifulSoup
import openpyxl

import datetime
from colorama import Fore

from random import randint
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, WebDriverException, TimeoutException
from selenium.webdriver.common.keys import Keys


class GeeInstaBot():
    def __init__(self, username, password):
        self.browserProfile = Options()
        # self.browserProfile.add_argument("--headless")  # Run in headless mode
        self.browserProfile.add_experimental_option('prefs', {'intl.accept_languages': 'en,en_US'})
        self.browserProfile.add_argument("--disable-blink-features=AutomationControlled") 
        self.browserProfile.add_experimental_option("excludeSwitches", ["enable-automation"]) 
        self.browserProfile.add_experimental_option("useAutomationExtension", False) 
        self.browser = webdriver.Chrome(options=self.browserProfile)
        self.username = username
        self.password = password
        self.cookies_file_path = "cookies_.pkl"
        self.max_retries = 15

        self.wait = WebDriverWait(self.browser, 10)

        self.dialog_profile_class_name = "x1dm5mii x16mil14 xiojian x1yutycm x1lliihq x193iq5w xh8yej3"

        self.error = Fore.RED
        self.success = Fore.GREEN
        self.info = Fore.CYAN

        self.errorIcon = "❌"
        self.successIcon = "✅"
        self.infoIcon = "ℹ️ "

        self.following = "Following"
        self.notFollowing = "Follow"

        self.filename = ""

    def progress_bar(self, progress, total):
        percent = 100 * (progress / float(total))
        bar = '◼︎' * int(percent) + '' * (100 - int(percent))
        print(Fore.YELLOW + f"\r|{bar}| {percent:.2f}%", end="\r")
        
    def save_cookies(self):
        cookies = self.browser.get_cookies()
        pickle.dump(cookies, open(self.cookies_file_path, "wb"))

    def generate_file_name(self, username, isFollowing=None):
        timestamp = int(time.time())
        if isFollowing == True:
            filename = f"{username}_followers_I'm_following{timestamp}.xlsx"
        elif isFollowing == False:
            filename = f"{username}_followers_I'm_not_following{timestamp}.xlsx"
        else:
            filename = f"{username}_followers_{timestamp}.xlsx"
        return filename
    
    def writeDataToExcel(self, data, username, isFollowing, headers = []):
        workbook = openpyxl.Workbook()

        worksheet = workbook.active
        headers = headers
        worksheet.append(headers)

        for row in data:
            worksheet.append(row)

        self.filename = self.generate_file_name(username, isFollowing)
        workbook.save(self.filename)

    def load_cookies(self):
        self.browser.get('https://www.instagram.com/')

        with open(self.cookies_file_path, 'rb') as f:
            cookies = pickle.load(f)

        # Check if any cookies are expired
        now = datetime.datetime.now()
        expired_cookies = []
        for cookie in cookies:
            expiration = cookie.get('expiry')
            if expiration:
                expiration_date = datetime.datetime.fromtimestamp(expiration)
                if expiration_date < now:
                    expired_cookies.append(cookie)
                    print("Cookie has expired:", cookie)
                    os.remove(self.cookies_file_path)
                    return

        if expired_cookies:
            print("[*] One or more cookies have expired. Deleting the cookies file and restarting the login process.")
            os.remove(self.cookies_file_path)
            self.signIn()
        else:
            # Add the cookies to the browser instance
            for cookie in cookies:
                self.browser.add_cookie(cookie)
        # Refresh the page to apply the cookies
        self.browser.refresh()

    def signIn(self):
        # Check if cookies file exists
        # 
        self.browser.set_window_size(1400, 920)
        retry_flag = False

        for retry in range(self.max_retries):
            try:
                if not os.path.exists(self.cookies_file_path):
                    self.browser.get('https://www.instagram.com/accounts/login/')
                    time.sleep(2)
                    usernameInput = self.browser.find_element('css selector', 'input[name="username"]')
                    passwordInput = self.browser.find_element('css selector', 'input[name="password"]')
                    usernameInput.send_keys(self.username)
                    passwordInput.send_keys(self.password)
                    passwordInput.send_keys(Keys.ENTER)
                    time.sleep(getRandomTime())

                    try:
                        """Close Notifications"""
                        self.browser.find_element(By.XPATH, '//button[contains(text(), "Not Now")]').click()
                    except NoSuchElementException:
                        pass

                    # Save cookies
                    self.save_cookies()
                else:
                    # if previous logged in Load cookies from file
                    self.load_cookies()

                time.sleep(getRandomTime())

            except WebDriverException as e:
                print(self.error + f'{self.errorIcon} Retry {retry + 1} failed: {str(e)}')
                retry_flag = True
            
            if not retry_flag:
                break

            if retry == self.max_retries - 1:
                print('[*] Maximum retries exceeded. Exiting.')

    
    def unfollowWithUsername(self, usernames = []):

        iter = 0
        self.progress_bar(iter, len(usernames))
        for username in usernames:
            self.browser.get('https://www.instagram.com/' + username + '/')
            iter += 1
            time.sleep(getRandomTime())

            try:
                """Close Notifications"""
                self.browser.find_element(By.XPATH, '//button[contains(text(), "Allow all cookies")]').click()
            except NoSuchElementException:
                pass

            time.sleep(getRandomTime())

            try:
                """Close Notifications"""
                self.browser.find_element(By.XPATH, '//button[contains(text(), "Not Now")]').click()
            except NoSuchElementException:
                pass

            followButton = self.browser.find_element(By.CSS_SELECTOR, 'button')
            self.progress_bar(iter, len(usernames))
            if followButton.text == 'Following':
                followButton.click()
                time.sleep(getRandomTime())
                confirmButton = self.browser.find_element(By.XPATH, '//span[text()="Unfollow"]')
                confirmButton.click()
                print(self.success + f"{self.successIcon} Unfollowed Successfully")
            else:
                print(self.info + f"{self.infoIcon} You are not following this user")

    def followWithUsername(self, usernames=[]):

        iter = 0
        self.progress_bar(iter, len(usernames))
        for username in usernames:
            self.browser.get('https://www.instagram.com/' + username + '/')
            time.sleep(getRandomTime())
            iter += 1
            # followButton = self.browser.find_element(By.CSS_SELECTOR, 'button')
            followButton = self.wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'button')))


            self.progress_bar(iter, len(usernames))
            if (followButton.text != 'Following'):
                followButton.click()
                time.sleep(getRandomTime())
                print(self.success + f"{self.successIcon} {username} Followed Successfully")
            else:
                print(self.info + f"{self.infoIcon} You are already following {username}")
            

    def getFollowersDetails(self, username):
        self.browser.get('https://www.instagram.com/' + username + '/')

        time.sleep(getRandomTime())
        
        ul_xpath = '/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[2]/div[2]/section/main/div/header/section/ul'

        try:
            ul_source = self.browser.find_element(By.XPATH, ul_xpath )
        except NoSuchElementException as e:
            print("[**] An Error Occured:", e)


        html = ul_source.get_attribute("innerHTML")

        soup = BeautifulSoup(html, 'html.parser')

        posts = soup.find('span', class_='_ac2a').get_text(strip=True)
        followers = soup.find('a', href=f'/{username}/followers/').find(class_='_ac2a').get('title')
        following = soup.find('a', href=f'/{username}/following/').find(class_='_ac2a').get_text(strip=True)

        print("Posts:", posts)
        print("Followers:", followers)
        print("Following:", following)

        return posts, followers, following


    def saveProfileToFile(self, followers_list, max_followers=None, username=None, isFollowing=None):
        followers = []
        fol = 0
        start_time = time.time()

        self.progress_bar(fol, len(followers_list))
        for follower in followers_list:
            fol += 1
                        
            try:
                name = follower.find(class_="x1lliihq x193iq5w x6ikm8r x10wlt62 xlyipyv xuxw1ft")
                user_name = follower.find(class_="x9f619 xjbqb8w x1rg5ohu x168nmei x13lgxp2 x5pf9jr xo71vjh x1n2onr6 x1plvlek xryxfnj x1c4vz4f x2lah0s x1q0g3np xqjyukv x6s0dn4 x1oa3qoh x1nhvcw1")
                img_src = follower.find("img")["src"]
                link_element = follower.find("a")["href"]
                status = follower.find(class_ = "_aacl _aaco _aacw _aad6 _aade")

                followers.append([
                    name.get_text(strip=True),
                    user_name.get_text(strip=True),
                    img_src,
                    f"https://instagram.com{link_element}",
                    status.get_text(strip=True)
                ])

                if (len(followers) == max_followers) or time.time() - start_time >= 2 and len(followers) == fol:
                    print(self.info + f"{self.infoIcon} No append in the last 2 seconds. Breaking...")
                    break

            except AttributeError:
                name = ""
                user_name = ""
                img_src = ""
                link_element = ""
                status = ""
            
            self.progress_bar(fol, len(followers_list))
            
        self.writeDataToExcel(followers, username, isFollowing, ["name", "username", "image", "profile"])
        print(self.info + f"{self.infoIcon} Followers extracted", fol)
        print(self.success + f"{self.successIcon} Successfully extracted and saved to file", fol)
        
        
    def getUserFollowers(self, username, max_followers, isFollowing=None):
        # followers url
        self.browser.get('https://www.instagram.com/' + username + '/followers/')
     
        time.sleep(getRandomTime())
        followers_ = "_aacl _aaco _aacw _aad6 _aade"

        retry_count = 0
        while True:
            try:
                time.sleep(3)
                pop_up_window = WebDriverWait(self.browser, 2).until(EC.element_to_be_clickable((By.CLASS_NAME, "_aano")))
                break  
            except TimeoutException:
                print(self.info + f"{self.infoIcon} Timeout occurred. Retrying...")
                retry_count += 1
        try:
            dialog_x_path = '/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[2]'
            followersDialog = self.browser.find_element(By.XPATH, dialog_x_path)   
            print(self.success + f"{self.successIcon} Dialog found..We're on!!")
        except NoSuchElementException:
            print(self.error + f"{self.errorIcon} Dialog xpath changed")    

        followersListNum = 0
        iteration = 0
        consecutive_empty_iterations = 0
        max_consecutive_empty_iterations = 5
        
        self.progress_bar(iteration, max_followers)
        while (followersListNum < max_followers):
            self.browser.execute_script('arguments[0].scrollTop = arguments[0].scrollTop + arguments[0].offsetHeight;', pop_up_window)

            # BeautifulSoup get html
            followersDialogHtml = followersDialog.get_attribute("innerHTML")
            soup = BeautifulSoup(followersDialogHtml, 'html.parser')
            followersList = soup.find_all(class_=self.dialog_profile_class_name)

            if len(followersList) < 1:
                print(self.error + f"{self.errorIcon} Dialog Profiles class name has definitely changed")
                break

            followers_list = []

            # Get followers of user that follow you or not
            if isFollowing is not None:
                for follower in followersList:
                    inner_div = follower.find(class_=followers_)
                    if inner_div is not None:
                        text = inner_div.get_text()
                        if (isFollowing == True and text == self.following) or (isFollowing == False and text == self.notFollowing):
                            followers_list.append(follower)


                new_followers = len(followers_list)
                
                iteration += 1
                self.progress_bar(iteration, new_followers)
                
                # if isFollowing == True:
                if new_followers == len(followers_list):
                    consecutive_empty_iterations += 1
                else:
                    consecutive_empty_iterations = 0
                
                if consecutive_empty_iterations == max_consecutive_empty_iterations:
                    print(self.info + f"{self.infoIcon} No new followers found.")
                    break
            
            # # if isFollowing is not passed, extract all followers
            else:
                followers_list += followersList
            
                new_followers = len(followers_list)

                if new_followers == 0:
                    consecutive_empty_iterations += 1
                else:
                    consecutive_empty_iterations = 0

                iteration += 1
                # print(self.info + f"{self.infoIcon} Extracting Profiles.... Please wait")
            
            self.progress_bar(iteration, len(followers_list))

            if consecutive_empty_iterations == max_consecutive_empty_iterations:
                print(self.info + f"{self.infoIcon} No new followers found.")
                break

            # Check if the maximum count is reached or if there are no more followers to extract
            if (max_followers is not None and len(followers_list) >= max_followers) or len(followers_list) == followersListNum:
                print(self.info + f"{self.infoIcon} Maximum number of followers reached or no more followers to extract.")
                break

        time.sleep(getRandomTime())
        # call function to save to excel
        self.saveProfileToFile(followers_list, max_followers, username, isFollowing)
        
    def getMyFollowing(self, max_followers):
        self.browser.get('https://www.instagram.com/' + self.username + '/following/')

        # get the following dialog
        time.sleep(getRandomTime())

        retry_count = 0

        while True:
            try:
                time.sleep(getRandomTime())
                pop_up_window = WebDriverWait(self.browser, 2).until(EC.element_to_be_clickable((By.CLASS_NAME, "_aano")))
                break  
            except TimeoutException:
                print(self.info + f"{self.infoIcon} Timeout occurred. Retrying...")
                retry_count += 1

        try:
            dialog_x_path = '/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[3]'
            followingDialog = self.browser.find_element(By.XPATH, dialog_x_path)
            print(self.success + f"{self.successIcon} Following Dialog Found")
        except NoSuchElementException:
            print(self.error + f"{self.errorIcon} Element x path changed")
        
        iteration = 0
        consecutive_empty_iterations = 0
        max_consecutive_empty_iterations = 100

        followersListNum = 0

        print(self.info + f"{self.infoIcon} Started Extraction...")
        self.progress_bar(iteration, max_followers)
        while (followersListNum < max_followers):
            self.browser.execute_script('arguments[0].scrollTop = arguments[0].scrollTop + arguments[0].offsetHeight;', pop_up_window)
            
            time.sleep(getRandomTime())

            followersDialogHtml = followingDialog.get_attribute("innerHTML")
            soup = BeautifulSoup(followersDialogHtml, 'html.parser')

            followersList = soup.find_all(class_=self.dialog_profile_class_name)
            if len(followersList) < 1:
                print(self.error + f"{self.errorIcon} Dialog Profiles class name has definitely changed")
                break
            
            followers_list = []
            followers_list += followersList
            
            new_followers = len(followers_list)

            if new_followers == 0:
                consecutive_empty_iterations += 1
            else:
                consecutive_empty_iterations = 0

            iteration += 1 
            self.progress_bar(iteration, len(followers_list))

            if consecutive_empty_iterations == max_consecutive_empty_iterations:
                print(self.info + f"{self.infoIcon} No new followers found.")
                break
            
            # Check if the maximum count is reached or if there are no more followers to extract
            if (len(followers_list) >= max_followers) or len(followers_list) == followersListNum:
                print(self.info + f"{self.infoIcon} Maximum number of followers reached or no more followers to extract.")
                break

        time.sleep(getRandomTime())
        self.saveProfileToFile(followers_list, self.username)

    
    def getUsernames(self, filename):
        workbook = openpyxl.load_workbook(filename=filename)

        # Select the active sheet
        sheet = workbook.active
        column_index = 2

        # Iterate through a specific column
        usernames = []
        for cell in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
            usernames.append(cell[0])
        return usernames

    def followProfiles(self):
        print(self.info + f"{self.infoIcon} Start following profiles")
        usernames = self.getUsernames(self.filename)
        self.followWithUsername(usernames)
    
    def unfollowProfiles(self, filename):
        usernames = self.getUsernames(filename)
        self.unfollowWithUsername(usernames)

def getRandomTime():
    randTime = randint(3, 5)
    return randTime

bot = GeeInstaBot('pabloe.gee', '@sventeen18!')
bot.signIn()
# bot.followWithUsername([])
bot.getUserFollowers('reneendunge', max_followers=50)
# bot.getMyFollowing(max_followers=50)
# bot.followProfiles()

